import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.drawing.image import Image
from io import BytesIO
from datetime import datetime
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import gdown
import os
import glob
import pytz
import re
import shutil
import config

# --- GESTIÓN DE ESTADO (SESSION STATE) ---
def inicializar_estado():
    defaults = {
        'pedidos': [],
        'carrito': [],
        'df_inventario_diario': None,
        'memoria_cliente': None,
        'memoria_fecha': datetime.today(),
        'memoria_busqueda_inv': "",
        'lista_revision': [],
        'reset_counter': 0,
        'info_archivo': None
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value

# --- CARGA DE DATOS ---
@st.cache_data
def cargar_catalogos():
    errores = []
    df_cli = pd.DataFrame()
    df_prod = pd.DataFrame()
    
    # 1. CLIENTES
    try:
        try: df_cli = pd.read_csv(config.FILE_CLIENTES, encoding='utf-8')
        except: df_cli = pd.read_csv(config.FILE_CLIENTES, encoding='latin-1')
        
        df_cli.columns = df_cli.columns.str.strip().str.upper()
        col_clave = next((c for c in df_cli.columns if 'CLAVE' in c or 'CODIGO' in c), df_cli.columns[0])
        col_nombre = next((c for c in df_cli.columns if ('CLIENTE' in c or 'NOMBRE' in c) and c != col_clave), df_cli.columns[1])
        
        df_cli = df_cli[[col_clave, col_nombre]].copy()
        df_cli.columns = ['CODIGO', 'NOMBRE']
        df_cli['DISPLAY'] = df_cli['CODIGO'].astype(str) + " - " + df_cli['NOMBRE'].astype(str)
    except Exception as e:
        errores.append(f"Clientes: {e}")

    # 2. PRODUCTOS
    try:
        try: df_prod = pd.read_csv(config.FILE_PRODUCTOS, encoding='utf-8')
        except: df_prod = pd.read_csv(config.FILE_PRODUCTOS, encoding='latin-1')
            
        df_prod.columns = df_prod.columns.str.strip().str.upper()
        col_clave = next(c for c in df_prod.columns if 'CLAVE' in c or 'CODIGO' in c)
        col_desc = next(c for c in df_prod.columns if 'NOMBRE' in c or 'DESCRIPCION' in c)
        col_sust = next((c for c in df_prod.columns if 'SUSTANCIA' in c), None)
        
        cols = [col_clave, col_desc]
        if col_sust: cols.append(col_sust)
        df_prod = df_prod[cols].copy()
        
        nombres = ['CODIGO', 'DESCRIPCION']
        if col_sust: nombres.append('SUSTANCIA')
        df_prod.columns = nombres
        
        if 'SUSTANCIA' not in df_prod.columns: df_prod['SUSTANCIA'] = '---'
        else: df_prod['SUSTANCIA'] = df_prod['SUSTANCIA'].fillna('---')

        # Limpieza
        df_prod['CODIGO'] = df_prod['CODIGO'].astype(str).str.strip()
        df_prod = df_prod.drop_duplicates(subset=['CODIGO'], keep='first')
        df_prod = df_prod.dropna(subset=['DESCRIPCION'])

        # Índice de búsqueda
        df_prod['SEARCH_INDEX'] = (
            df_prod['CODIGO'] + " | " + 
            df_prod['DESCRIPCION'].astype(str) + " | " + 
            df_prod['SUSTANCIA'].astype(str)
        ).str.upper()
        
    except Exception as e:
        errores.append(f"Productos: {e}")

    return df_cli, df_prod, errores

# --- PROCESAMIENTO DE INVENTARIO ---
def procesar_inventario(df_raw, df_productos):
    df_tj = df_raw.iloc[:, [0, 1, 5, 6]].copy()
    df_tj.columns = ['CODIGO', 'PRODUCTO', 'CORTA_CAD', 'EXISTENCIA']
    df_tj = df_tj.dropna(subset=['CODIGO'])
    df_tj['CODIGO'] = df_tj['CODIGO'].astype(str).str.strip()

    df_merged = pd.merge(df_tj, df_productos[['CODIGO', 'SUSTANCIA']], on='CODIGO', how='left')
    df_merged['SUSTANCIA'] = df_merged['SUSTANCIA'].fillna('---')
    
    df_merged['INDICE_BUSQUEDA'] = (
        df_merged['CODIGO'] + " " + 
        df_merged['PRODUCTO'] + " " + 
        df_merged['SUSTANCIA']
    ).str.upper()

    cols_finales = ['CODIGO', 'PRODUCTO', 'SUSTANCIA', 'EXISTENCIA', 'CORTA_CAD', 'INDICE_BUSQUEDA']
    return df_merged[cols_finales]

# --- GOOGLE DRIVE ---
@st.cache_data(ttl=600, show_spinner=False)
def descargar_de_drive(folder_id):
    try:
        url = f'https://drive.google.com/drive/folders/{folder_id}'
        output_dir = config.TEMP_DRIVE_FOLDER
        
        if os.path.exists(output_dir):
            shutil.rmtree(output_dir)
        os.makedirs(output_dir, exist_ok=True)
        
        gdown.download_folder(url, output=output_dir, quiet=True, use_cookies=False)
        
        archivos = glob.glob(f"{output_dir}/*.xlsx") + glob.glob(f"{output_dir}/*.csv")
        
        if archivos:
            def puntaje_novedad(ruta_archivo):
                mtime = os.path.getmtime(ruta_archivo)
                nombre = os.path.basename(ruta_archivo)
                match = re.search(r'\((\d+)\)\.', nombre)
                version = int(match.group(1)) if match else 0
                return (mtime, version)

            archivo_mas_reciente = max(archivos, key=puntaje_novedad)
            
            timestamp = os.path.getmtime(archivo_mas_reciente)
            dt_utc = datetime.fromtimestamp(timestamp, pytz.utc)
            zona_local = pytz.timezone(config.TIMEZONE)
            dt_local = dt_utc.astimezone(zona_local)
            fecha_mod = dt_local.strftime('%d/%m/%Y %H:%M')
            
            nombre_archivo = os.path.basename(archivo_mas_reciente)
            
            if archivo_mas_reciente.endswith('.csv'):
                try: df = pd.read_csv(archivo_mas_reciente, header=1, encoding='latin-1')
                except: df = pd.read_csv(archivo_mas_reciente, header=1, encoding='utf-8')
            else:
                df = pd.read_excel(archivo_mas_reciente, header=1)
                
            return df, nombre_archivo, fecha_mod
        else:
            return None, None, None
            
    except Exception as e:
        return None, f"Error: {str(e)}", None

# --- GENERACIÓN DE IMAGEN (MATPLOTLIB) ---
def generar_imagen_lista(df_rev, cliente_foto, incluir_sustancia):
    df_plot = df_rev.copy()
    if not incluir_sustancia and 'SUSTANCIA' in df_plot.columns:
        df_plot = df_plot.drop(columns=['SUSTANCIA'])
            
    cell_colors = []
    hay_rojo = False
    hay_amarillo = False
    
    for _, row in df_plot.iterrows():
        ex = pd.to_numeric(row['EXISTENCIA'], errors='coerce') or 0
        cc = pd.to_numeric(row['CORTA_CAD'], errors='coerce') or 0
        
        if ex == 0 and cc == 0:
            fila_color = ['#fe9292'] * len(df_plot.columns)
            hay_rojo = True
        elif ex == 0 and cc > 0:
            fila_color = ['#ffe59a'] * len(df_plot.columns)
            hay_amarillo = True
        else:
            fila_color = ['#ffffff'] * len(df_plot.columns)
        cell_colors.append(fila_color)

    num_filas = len(df_plot)
    num_cols = len(df_plot.columns)
    ancho_dinamico = max(10, num_cols * 1.25) 
    alto_dinamico = num_filas * 0.35
    
    if cliente_foto: alto_dinamico += 0.2
    if hay_rojo or hay_amarillo: alto_dinamico += 0.2
    
    fig, ax = plt.subplots(figsize=(ancho_dinamico, alto_dinamico)) 
    ax.axis('off')
    
    if cliente_foto:
        cod, nom = cliente_foto.split(" - ", 1)
        plt.title(f"{nom}\n{cod}", fontsize=16, fontweight='bold', pad=20)

    tabla = ax.table(
        cellText=df_plot.values,
        colLabels=df_plot.columns,
        cellColours=cell_colors,
        cellLoc='center',
        loc='center'
    )
    
    tabla.auto_set_font_size(False)
    tabla.set_fontsize(11)
    tabla.scale(1, 1.5)
    tabla.auto_set_column_width(col=list(range(len(df_plot.columns))))
    
    leyendas = []
    if hay_amarillo: leyendas.append(mpatches.Patch(color='#ffe59a', label='SOLO CORTA CAD.'))
    if hay_rojo: leyendas.append(mpatches.Patch(color='#fe9292', label='NO DISPONIBLE'))
        
    if leyendas:
        plt.legend(handles=leyendas, loc='upper center', bbox_to_anchor=(0.5, -0.02), ncol=2, frameon=False, fontsize=10)

    buf = BytesIO()
    plt.savefig(buf, format='png', bbox_inches='tight', dpi=150, pad_inches=0.5)
    buf.seek(0)
    plt.close(fig) # Importante cerrar la figura
    return buf

# --- GENERACIÓN DE EXCEL (OPENPYXL) ---
def generar_excel_pedidos(pedidos_lista):
    try:
        wb = openpyxl.load_workbook(config.FILE_PLANTILLA)
        base = wb.active; base.title = "Base"
        conteo = {}
        
        for p in pedidos_lista:
            cod = p['cli_cod']
            conteo[cod] = conteo.get(cod, 0) + 1
            nom_hoja = cod if conteo[cod] == 1 else f"{cod}-{conteo[cod]}"
            ws = wb.copy_worksheet(base); ws.title = nom_hoja
            
            ws['B2'] = "SUC. TIJ"; ws['B3'] = "LUIS FELIPE GARCÍA DOMÍNGUEZ"
            ws['B4'] = p['cli_nom']; ws['D6'] = p['fecha'].strftime('%d/%m/%Y')
            try: ws['B6'] = int(cod)
            except: ws['B6'] = cod
            
            try:
                img = Image(config.FILE_IMAGEN)
                img.width = 270; img.height = 80; img.anchor = 'D1'
                ws.add_image(img)
            except: pass
            
            datos = p['items'][['CODIGO', 'DESCRIPCION', 'SOLICITADA', 'SURTIDO', 'O.C.']].values.tolist()
            for idx, row in enumerate(datos):
                for c, val in enumerate(row): ws.cell(row=10+idx, column=c+1, value=val)
        
        del wb['Base']
        b = BytesIO(); wb.save(b); b.seek(0)
        return b
    except Exception as e:
        return None
