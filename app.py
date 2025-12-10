import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.drawing.image import Image
from io import BytesIO
from datetime import datetime
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import gdown
import os
import glob
import pytz

# --- CONFIGURACIÓN DE PÁGINA ---
st.set_page_config(page_title="Sistema Ventas", page_icon="💊", layout="wide")

# --- ARCHIVOS EN GITHUB ---
FILE_CLIENTES = 'clientes.csv'
FILE_PRODUCTOS = 'productos.csv'
FILE_PLANTILLA = 'plantilla.xlsx'
FILE_IMAGEN = 'logo.png'

# --- ESTADO DE LA APP (MEMORIA) ---
# Faltantes
if 'pedidos' not in st.session_state: st.session_state.pedidos = []
if 'carrito' not in st.session_state: st.session_state.carrito = []
# Inventario (Nueva variable para persistencia)
if 'df_inventario_diario' not in st.session_state: st.session_state.df_inventario_diario = None
# --- NUEVO: PERSISTENCIA DE DATOS (CLIENTE Y FECHA) ---
if 'memoria_cliente' not in st.session_state: st.session_state.memoria_cliente = None
if 'memoria_fecha' not in st.session_state: st.session_state.memoria_fecha = datetime.today()

# --- FUNCIÓN DE CARGA DE DATOS (CATÁLOGOS MAESTROS) ---
@st.cache_data
def cargar_catalogos():
    errores = []
    df_cli = pd.DataFrame()
    df_prod = pd.DataFrame()
    
    # 1. CLIENTES
    try:
        try: df_cli = pd.read_csv(FILE_CLIENTES, encoding='utf-8')
        except: df_cli = pd.read_csv(FILE_CLIENTES, encoding='latin-1')
        
        df_cli.columns = df_cli.columns.str.strip().str.upper()
        col_clave = next((c for c in df_cli.columns if 'CLAVE' in c or 'CODIGO' in c), df_cli.columns[0])
        col_nombre = next((c for c in df_cli.columns if ('CLIENTE' in c or 'NOMBRE' in c) and c != col_clave), df_cli.columns[1])
        
        df_cli = df_cli[[col_clave, col_nombre]].copy()
        df_cli.columns = ['CODIGO', 'NOMBRE']
        df_cli['DISPLAY'] = df_cli['CODIGO'].astype(str) + " - " + df_cli['NOMBRE'].astype(str)
    except Exception as e:
        errores.append(f"Clientes: {e}")

    # 2. PRODUCTOS (Catálogo Maestro)
    try:
        try: df_prod = pd.read_csv(FILE_PRODUCTOS, encoding='utf-8')
        except: df_prod = pd.read_csv(FILE_PRODUCTOS, encoding='latin-1')
            
        df_prod.columns = df_prod.columns.str.strip().str.upper()
        col_clave = next(c for c in df_prod.columns if 'CLAVE' in c or 'CODIGO' in c)
        col_desc = next(c for c in df_prod.columns if 'NOMBRE' in c or 'DESCRIPCION' in c)
        col_sust = next((c for c in df_prod.columns if 'SUSTANCIA' in c), None)
        
        cols = [col_clave, col_desc]
        if col_sust: cols.append(col_sust)
        df_prod = df_prod[cols].copy()
        
        nombres = ['CODIGO', 'DESCRIPCION']
        if col_sust: nombres.append('SUSTANCIA')
        df_prod.columns = nombres
        
        if 'SUSTANCIA' not in df_prod.columns: df_prod['SUSTANCIA'] = '---'
        else: df_prod['SUSTANCIA'] = df_prod['SUSTANCIA'].fillna('---')

        # --- LIMPIEZA AGRESIVA (NUEVO) ---
        # 1. Asegurar que el código sea texto limpio (sin espacios invisibles)
        df_prod['CODIGO'] = df_prod['CODIGO'].astype(str).str.strip()
        
        # 2. Eliminar códigos duplicados (Se queda con la primera aparición)
        # Esto soluciona que te salgan "varias filas" si el código se repite en el archivo
        df_prod = df_prod.drop_duplicates(subset=['CODIGO'], keep='first')
        
        # 3. Eliminar productos sin nombre
        df_prod = df_prod.dropna(subset=['DESCRIPCION'])
        # ---------------------------------

        # Índice de búsqueda optimizado
        df_prod['SEARCH_INDEX'] = (
            df_prod['CODIGO'] + " | " + 
            df_prod['DESCRIPCION'].astype(str) + " | " + 
            df_prod['SUSTANCIA'].astype(str)
        ).str.upper()
        
    except Exception as e:
        errores.append(f"Productos: {e}")

    return df_cli, df_prod, errores

# Cargar datos al inicio
df_clientes, df_productos, logs = cargar_catalogos()

# --- NAVEGACIÓN LATERAL ---
with st.sidebar:
    st.title("Navegación")
    vista = st.radio("Ir a:", ["🔍 Revisar Existencias", "📝 Reportar Faltantes"])
    st.divider()
    
    # Indicadores de estado
    st.caption("Estado del Sistema:")
    if st.session_state.df_inventario_diario is not None:
        st.success("✅ Inventario Diario Cargado")
    else:
        st.warning("⚠️ Falta Inventario Diario")
        
    if logs:
        for l in logs: st.error(l)
            
# ==============================================================================
# VISTA 1: REVISAR EXISTENCIAS (INVENTARIO DIARIO)
# ==============================================================================

if vista == "🔍 Revisar Existencias":
    st.header("🔍 Buscador de Existencias")

    # --- CONFIGURACIÓN DE CARPETA DRIVE ---
    # Pega aquí el ID de tu CARPETA PÚBLICA (lo que sigue de folders/...)
    DRIVE_FOLDER_ID = "1bvF7yuIRiJQ0oiXiZ6s3JD8goy1DUi1K"  # <--- ¡PÉGALO AQUÍ!

    # --- ESTADOS INICIALES ---
    if 'lista_revision' not in st.session_state: st.session_state.lista_revision = []
    if 'reset_counter' not in st.session_state: st.session_state.reset_counter = 0

    # --- FUNCIÓN: DESCARGAR CARPETA (Con Ajuste de Zona Horaria) ---
    @st.cache_data(ttl=600, show_spinner=False)
    def descargar_de_drive(folder_id):
        try:
            import pytz # Librería para zonas horarias
            
            url = f'https://drive.google.com/drive/folders/{folder_id}'
            output_dir = './temp_drive_folder'
            
            if os.path.exists(output_dir):
                import shutil
                shutil.rmtree(output_dir)
            os.makedirs(output_dir, exist_ok=True)
            
            gdown.download_folder(url, output=output_dir, quiet=True, use_cookies=False)
            
            archivos = glob.glob(f"{output_dir}/*.xlsx") + glob.glob(f"{output_dir}/*.csv")
            
            if archivos:
                import re
                
                def puntaje_novedad(ruta_archivo):
                    mtime = os.path.getmtime(ruta_archivo)
                    nombre = os.path.basename(ruta_archivo)
                    match = re.search(r'\((\d+)\)\.', nombre)
                    version = int(match.group(1)) if match else 0
                    return (mtime, version)

                archivo_mas_reciente = max(archivos, key=puntaje_novedad)
                
                # --- CORRECCIÓN DE HORA AQUÍ ---
                timestamp = os.path.getmtime(archivo_mas_reciente)
                
                # 1. Interpretamos la fecha como UTC (Hora del servidor)
                dt_utc = datetime.fromtimestamp(timestamp, pytz.utc)
                
                # 2. Convertimos a hora de TIJUANA
                zona_tijuana = pytz.timezone('America/Tijuana')
                dt_local = dt_utc.astimezone(zona_tijuana)
                
                fecha_mod = dt_local.strftime('%d/%m/%Y %H:%M')
                # -------------------------------
                
                nombre_archivo = os.path.basename(archivo_mas_reciente)
                
                if archivo_mas_reciente.endswith('.csv'):
                    try: df = pd.read_csv(archivo_mas_reciente, header=1, encoding='latin-1')
                    except: df = pd.read_csv(archivo_mas_reciente, header=1, encoding='utf-8')
                else:
                    df = pd.read_excel(archivo_mas_reciente, header=1)
                    
                return df, nombre_archivo, fecha_mod
            else:
                return None, None, None
                
        except Exception as e:
            return None, f"Error: {str(e)}", None

    # --- FUNCIÓN: PROCESAR DATA ---
    def procesar_inventario(df_raw):
        df_tj = df_raw.iloc[:, [0, 1, 5, 6]].copy()
        df_tj.columns = ['CODIGO', 'PRODUCTO', 'CORTA_CAD', 'EXISTENCIA']
        df_tj = df_tj.dropna(subset=['CODIGO'])
        df_tj['CODIGO'] = df_tj['CODIGO'].astype(str).str.strip()

        df_merged = pd.merge(df_tj, df_productos[['CODIGO', 'SUSTANCIA']], on='CODIGO', how='left')
        df_merged['SUSTANCIA'] = df_merged['SUSTANCIA'].fillna('---')
        
        df_merged['INDICE_BUSQUEDA'] = (
            df_merged['CODIGO'] + " " + 
            df_merged['PRODUCTO'] + " " + 
            df_merged['SUSTANCIA']
        ).str.upper()

        cols_finales = ['CODIGO', 'PRODUCTO', 'SUSTANCIA', 'EXISTENCIA', 'CORTA_CAD', 'INDICE_BUSQUEDA']
        return df_merged[cols_finales]

    # --- LÓGICA DE CARGA ---
    uploaded_file = st.file_uploader("📤 Cargar archivo local (sobrescribe)", type=['csv', 'xlsx'])
    
    df_activo = None
    info_origen = ""

    # CASO A: Local
    if uploaded_file:
        try:
            if uploaded_file.name.endswith('.csv'):
                try: df_raw = pd.read_csv(uploaded_file, header=1, encoding='latin-1')
                except: uploaded_file.seek(0); df_raw = pd.read_csv(uploaded_file, header=1, encoding='utf-8')
            else:
                df_raw = pd.read_excel(uploaded_file, header=1)
            
            df_activo = procesar_inventario(df_raw)
            # Guardamos todo en sesión
            st.session_state.df_inventario_diario = df_activo
            st.session_state.info_archivo = f"Local: {uploaded_file.name}"
            info_origen = st.session_state.info_archivo
            
        except Exception as e:
            st.error(f"Error archivo local: {e}")

    # CASO B: Memoria (Ya cargado)
    elif st.session_state.df_inventario_diario is not None:
        df_activo = st.session_state.df_inventario_diario
        # Recuperamos la info del archivo guardada
        info_origen = st.session_state.get('info_archivo', 'Memoria')

    # CASO C: Carpeta Drive (Automático)
    elif DRIVE_FOLDER_ID:
        with st.spinner("☁️ Sincronizando con Drive (esto toma unos segundos)..."):
            # Llamamos a la función cacheada
            df_cloud, nombre_archivo, fecha_mod = descargar_de_drive(DRIVE_FOLDER_ID)
            
            if df_cloud is not None:
                df_activo = procesar_inventario(df_cloud)
                
                # Guardamos en sesión
                st.session_state.df_inventario_diario = df_activo
                
                # Creamos el texto de información
                info_str = f"☁️ Nube: {nombre_archivo} | 📅 Fecha: {fecha_mod}"
                st.session_state.info_archivo = info_str
                info_origen = info_str
                
                # Rerun para mostrar los datos inmediatamente
                st.rerun()
            else:
                # Si nombre_archivo tiene texto, es que trajo un error
                if nombre_archivo and "Error" in nombre_archivo:
                    st.error(nombre_archivo)
                else:
                    st.warning("⚠️ Carpeta vacía o sin acceso.")

    # --- RENDERIZADO ---
    if df_activo is not None:
        # Mostrar barra de info con versión
        st.success(f"✅ {info_origen}")
        
        col_search, col_reset = st.columns([4, 1])
        with col_reset:
            # Botón Recargar: Limpia la memoria Y la cache de descarga
            if st.button("🔄 Recargar Nube"):
                st.session_state.df_inventario_diario = None
                descargar_de_drive.clear() # Limpia la cache de la función de descarga
                st.rerun()
        
        busqueda = st.text_input("¿Qué buscas?", placeholder="Nombre, Clave o Sustancia...").upper()
        
        resultados = pd.DataFrame()
        
        if busqueda:
            mask = df_activo['INDICE_BUSQUEDA'].str.contains(busqueda, na=False)
            resultados = df_activo[mask].drop(columns=['INDICE_BUSQUEDA'])
            st.success(f"Encontrados: {len(resultados)}")
            
            dynamic_key = f"search_table_{st.session_state.reset_counter}"
            
            event = st.dataframe(
                resultados,
                width="stretch",
                hide_index=True,
                on_select="rerun", 
                selection_mode="multi-row",
                key=dynamic_key 
            )
            
            if len(event.selection.rows) > 0:
                st.divider()
                # --- NUEVO: COLUMNAS PARA BOTÓN Y CANTIDAD ---
                c_btn, c_qty = st.columns([3, 1])
                
                # Input de cantidad (opcional, por defecto 0)
                qty_add = c_qty.number_input("Piezas (Opcional):", min_value=0, value=0, key="qty_add_rev")
                
                if c_btn.button(f"⬇️ Agregar Selección ({len(event.selection.rows)})"):
                    filas_seleccionadas = resultados.iloc[event.selection.rows].copy()
                    
                    # Agregar columna de piezas
                    # Si es 0, mostramos "-", si tiene número, lo mostramos.
                    filas_seleccionadas['SOLICITADO'] = qty_add if qty_add > 0 else "-"
                    
                    nuevos_items = filas_seleccionadas.to_dict('records')
                    st.session_state.lista_revision.extend(nuevos_items)
                    
                    st.session_state.reset_counter += 1 
                    st.toast("✅ Agregado")
                    st.rerun() 
        else:
            st.info("Inventario cargado. Escribe arriba para filtrar.")

        # --- SECCIÓN INFERIOR: TABLA DE REVISIÓN ACUMULADA ---
        st.divider()
        st.subheader("📋 Tu Lista de Revisión")
        
        # Columnas para los botones de acción
        col_info, col_borrar_sel, col_borrar_todo = st.columns([3, 2, 1])
        
        if st.session_state.lista_revision:
            df_rev = pd.DataFrame(st.session_state.lista_revision)
            
            # Orden de columnas
            cols_orden = ['CODIGO', 'PRODUCTO', 'SUSTANCIA', 'EXISTENCIA', 'CORTA_CAD', 'SOLICITADO']
            for c in cols_orden:
                if c not in df_rev.columns: df_rev[c] = "-"
            df_rev = df_rev[cols_orden]

            # Estilos
            def estilo_existencias(row):
                existencia = pd.to_numeric(row['EXISTENCIA'], errors='coerce') or 0
                corta_cad = pd.to_numeric(row['CORTA_CAD'], errors='coerce') or 0
                colores = [''] * len(row)
                if existencia == 0 and corta_cad == 0:
                    colores = ['background-color: #390D10'] * len(row)
                elif existencia == 0 and corta_cad > 0:
                    colores = ['background-color: #4B3718'] * len(row)
                return colores

            # --- TABLA INTERACTIVA (CON SELECCIÓN) ---
            # Guardamos el evento para saber qué filas seleccionaste
            event_revision = st.dataframe(
                df_rev.style.apply(estilo_existencias, axis=1),
                width="stretch",
                hide_index=True,
                on_select="rerun",          # <--- Activamos selección
                selection_mode="multi-row", # <--- Selección múltiple
                key="tabla_revision_final"
            )
            
            # --- LÓGICA DE BORRADO SELECTIVO ---
            filas_seleccionadas = event_revision.selection.rows
            
            with col_borrar_sel:
                # El botón solo aparece si seleccionaste algo
                if filas_seleccionadas:
                    if st.button(f"🗑️ Borrar ({len(filas_seleccionadas)}) seleccionados"):
                        # Reconstruimos la lista EXCLUYENDO los índices seleccionados
                        indices_a_borrar = set(filas_seleccionadas)
                        st.session_state.lista_revision = [
                            item for i, item in enumerate(st.session_state.lista_revision) 
                            if i not in indices_a_borrar
                        ]
                        st.rerun()

            with col_borrar_todo:
                if st.button("🔥 Borrar Todo"):
                    st.session_state.lista_revision = []
                    st.rerun()

            # --- CONFIGURACIÓN DE IMAGEN ---
            st.divider()
            st.caption("Configuración de la Imagen:")
            
            c_cli, c_opt = st.columns([2, 1])
            with c_cli:
                cliente_foto = st.selectbox("Título de Cliente (Opcional):", options=df_clientes['DISPLAY'], index=None, placeholder="Sin título...", key="cli_foto_input")
            with c_opt:
                incluir_sustancia = st.checkbox("Incluir columna 'Sustancia'", value=True)

            if st.button("📸 Descargar Tabla como Imagen"):
                try:
                    # 1. FILTRAR DATOS
                    df_plot = df_rev.copy()
                    
                    if not incluir_sustancia:
                        if 'SUSTANCIA' in df_plot.columns:
                            df_plot = df_plot.drop(columns=['SUSTANCIA'])
                            
                    # 2. COLORES
                    cell_colors = []
                    hay_rojo = False
                    hay_amarillo = False
                    
                    for _, row in df_plot.iterrows():
                        ex = pd.to_numeric(row['EXISTENCIA'], errors='coerce') or 0
                        cc = pd.to_numeric(row['CORTA_CAD'], errors='coerce') or 0
                        
                        if ex == 0 and cc == 0:
                            fila_color = ['#fe9292'] * len(df_plot.columns)
                            hay_rojo = True
                        elif ex == 0 and cc > 0:
                            fila_color = ['#ffe59a'] * len(df_plot.columns)
                            hay_amarillo = True
                        else:
                            fila_color = ['#ffffff'] * len(df_plot.columns)
                        cell_colors.append(fila_color)

                    # 3. DIMENSIONES DINÁMICAS (ANCHO Y ALTO VARIABLE)
                    num_filas = len(df_plot)
                    num_cols = len(df_plot.columns)
                    
                    # Ancho: 2.5 pulgadas por columna (aprox) para dar buen espacio al texto
                    ancho_dinamico = max(10, num_cols * 1.25) 
                    
                    # Alto: 0.5 pulgadas por fila + espacio extra para encabezados/títulos
                    alto_dinamico = num_filas * 0.35
                    
                    if cliente_foto: alto_dinamico += 0.2
                    if hay_rojo or hay_amarillo: alto_dinamico += 0.2
                    
                    fig, ax = plt.subplots(figsize=(ancho_dinamico, alto_dinamico)) 
                    ax.axis('off')
                    
                    # 4. TÍTULO
                    if cliente_foto:
                        cod, nom = cliente_foto.split(" - ", 1)
                        # pad=20 da un pequeño aire interno antes del margen blanco
                        plt.title(f"{nom}\n{cod}", fontsize=16, fontweight='bold', pad=20)

                    # 5. DIBUJAR TABLA
                    tabla = ax.table(
                        cellText=df_plot.values,
                        colLabels=df_plot.columns,
                        cellColours=cell_colors,
                        cellLoc='center',
                        loc='center'
                    )
                    
                    # Estilizado
                    tabla.auto_set_font_size(False)
                    tabla.set_fontsize(11)
                    tabla.scale(1, 1.5) # Celdas más altas para mejor lectura
                    tabla.auto_set_column_width(col=list(range(len(df_plot.columns))))
                    
                    # 6. LEYENDA
                    leyendas = []
                    if hay_amarillo:
                        leyendas.append(mpatches.Patch(color='#ffe59a', label='SOLO CORTA CAD.'))
                    if hay_rojo:
                        leyendas.append(mpatches.Patch(color='#fe9292', label='NO DISPONIBLE'))
                        
                    if leyendas:
                        plt.legend(
                            handles=leyendas, 
                            loc='upper center', 
                            bbox_to_anchor=(0.5, -0.02), 
                            ncol=2, 
                            frameon=False,
                            fontsize=10
                        )

                    # Guardar
                    buf = BytesIO()
                    
                    # --- AQUÍ ESTÁ LA MAGIA DE LOS MÁRGENES ---
                    # bbox_inches='tight': Recorta todo el lienzo sobrante ajustándose al contenido exacto.
                    # pad_inches=0.5: Agrega exactamente 1/2 pulgada de margen blanco ALREDEDOR de ese recorte.
                    plt.savefig(buf, format='png', bbox_inches='tight', dpi=150, pad_inches=0.5)
                    
                    buf.seek(0)
                    
                    st.download_button(
                        label="⬇️ Guardar PNG",
                        data=buf,
                        file_name="Lista_Revision.png",
                        mime="image/png"
                    )
                except Exception as e:
                    st.error(f"Error generando imagen: {e}")

        else:
            st.caption("Selecciona productos arriba para armar tu lista de revisión.")

# ==============================================================================
# VISTA 2: REPORTAR FALTANTES (POS)
# ==============================================================================
elif vista == "📝 Reportar Faltantes":
    st.header("📝 Generador de Reporte de Faltantes")
    
    # --- BOTÓN DE REINICIO EN LA BARRA LATERAL ---
    with st.sidebar:
        st.divider()
        st.markdown("### ⚙️ Acciones")
        # Usamos type="primary" para que salga rojo/destacado
        if st.button("🗑️ BORRAR TODO (Reiniciar)", type="primary"):
            st.session_state.pedidos = []
            st.session_state.carrito = []
            st.session_state.cliente_box = None
            st.session_state.memoria_cliente = None # <--- NUEVO: Limpiar memoria
            st.rerun()
    # ----------------------------------------------------
    
    # Callbacks
    def agregar_producto():
        cliente = st.session_state.cliente_box
        prod_str = st.session_state.prod_box
        cant = st.session_state.qty_box
        
        if cliente and prod_str:
            row = df_productos[df_productos['SEARCH_INDEX'] == prod_str].iloc[0]
            item = {
                "CODIGO": row['CODIGO'],
                "DESCRIPCION": row['DESCRIPCION'],
                "SOLICITADA": cant,
                "SURTIDO": 0,
                "O.C.": "N/A"
            }
            st.session_state.carrito.append(item)
            st.session_state.qty_box = 1      
            st.session_state.prod_box = None 
        else:
            st.warning("⚠️ Selecciona Cliente y Producto")

    def finalizar_pedido_cb():
        if st.session_state.cliente_box:
            # ... (código de guardado del pedido) ...
                        
            st.session_state.pedidos.append(pedido_nuevo)
            st.session_state.carrito = []
            st.session_state.cliente_box = None
            st.session_state.memoria_cliente = None # <--- NUEVO: Limpiar memoria
            st.session_state.search_faltantes_input = "" 
        else:
            st.error("Falta Cliente")

    tab1, tab2 = st.tabs(["1. Registrar", "2. Descargar Excel"])
    
    with tab1:
        col1, col2 = st.columns([1, 2])
        
        # --- COLUMNA IZQUIERDA: BÚSQUEDA Y SELECCIÓN ---
        with col1:
            st.subheader("Datos")
            
            # --- LÓGICA DE PERSISTENCIA PARA CLIENTE ---
            # 1. Calculamos el índice donde está el cliente guardado
            lista_opciones = df_clientes['DISPLAY'].tolist()
            try:
                idx_guardado = lista_opciones.index(st.session_state.memoria_cliente)
            except:
                idx_guardado = None

            # 2. Función para actualizar la memoria cuando cambies el cliente
            def actualizar_cliente():
                st.session_state.memoria_cliente = st.session_state.cliente_box

            st.selectbox(
                "Cliente:", 
                options=df_clientes['DISPLAY'], 
                index=idx_guardado, # Usamos el índice recuperado
                placeholder="Buscar...", 
                key="cliente_box", 
                on_change=actualizar_cliente # Guardamos cambios al momento
            )
            
            # --- LÓGICA DE PERSISTENCIA PARA FECHA ---
            def actualizar_fecha():
                st.session_state.memoria_fecha = st.session_state.fecha_box

            fecha_input = st.date_input(
                "Fecha:", 
                value=st.session_state.memoria_fecha, # Usamos valor recuperado
                key="fecha_box", # Cambié el key para diferenciarlo
                on_change=actualizar_fecha
            )
            
            st.divider()
            st.subheader("Producto")
            
            # 1. Input de Búsqueda (Texto)
            query_faltantes = st.text_input("Buscar:", placeholder="Nombre, Clave o Sustancia...", key="search_faltantes_input").upper()
            
            # Inicializar contador para resetear la tabla de búsqueda
            if 'reset_search_faltantes' not in st.session_state:
                st.session_state.reset_search_faltantes = 0
            
            if query_faltantes:
                # 2. Filtrar Resultados (Busca en el índice sucio pero completo)
                mask = df_productos['SEARCH_INDEX'].str.contains(query_faltantes, na=False)
                resultados_f = df_productos[mask].copy()
                
                # --- LÓGICA DE LIMPIEZA "VISTA 1" ---
                
                # A. Eliminar vacíos en Descripción
                resultados_f = resultados_f.dropna(subset=['DESCRIPCION'])
                
                # B. ELIMINAR DUPLICADOS POR CÓDIGO (La clave para que se vea limpio)
                # Esto fuerza a que solo exista 1 fila por cada código único.
                resultados_f = resultados_f.drop_duplicates(subset=['CODIGO'], keep='first')
                
                # C. Seleccionar solo las columnas bonitas (Ocultamos el índice de búsqueda)
                # Aseguramos el orden: CÓDIGO | DESCRIPCION | SUSTANCIA
                cols_mostrar = ['CODIGO', 'DESCRIPCION', 'SUSTANCIA']
                # Filtramos solo las columnas que realmente existen para evitar errores
                cols_existentes = [c for c in cols_mostrar if c in resultados_f.columns]
                resultados_f = resultados_f[cols_existentes]
                
                # -------------------------------------
                
               # 3. Mostrar Tabla para Seleccionar
                key_table = f"table_results_{st.session_state.reset_search_faltantes}"
                
                event_f = st.dataframe(
                    resultados_f, 
                    width="stretch",
                    hide_index=True,
                    on_select="rerun",
                    selection_mode="single-row", 
                    key=key_table
                    # HE BORRADO LA LÍNEA: height=200 
                    # Al borrarla, la tabla se encoge automáticamente al tamaño del contenido.
                )
                
                # 4. Si hay selección, mostramos controles de agregar
                if len(event_f.selection.rows) > 0:
                    idx = event_f.selection.rows[0]
                    row_selected = resultados_f.iloc[idx]
                    
                    st.success(f"Seleccionado: **{row_selected['DESCRIPCION']}**")
                    
                    c_qty, c_btn = st.columns([1, 1])
                    cantidad = c_qty.number_input("Cantidad:", min_value=1, value=1, key="qty_faltantes_input")
                    
                    def agregar_seleccion():
                        if st.session_state.cliente_box:
                            item = {
                                "CODIGO": row_selected['CODIGO'],
                                "DESCRIPCION": row_selected['DESCRIPCION'],
                                "SOLICITADA": cantidad,
                                "SURTIDO": 0,
                                "O.C.": "N/A"
                            }
                            st.session_state.carrito.append(item)
                            
                            st.session_state.reset_search_faltantes += 1 
                            st.session_state.search_faltantes_input = "" 
                            st.session_state.qty_faltantes_input = 1     
                        else:
                            st.warning("⚠️ ¡Falta seleccionar el Cliente arriba!")

                    c_btn.button("➕ Agregar", on_click=agregar_seleccion, use_container_width=True)

        # --- COLUMNA DERECHA: CARRITO (Igual que antes) ---
        with col2:
            st.subheader("🛒 Carrito")
            if st.session_state.carrito:
                df_cart = pd.DataFrame(st.session_state.carrito)
                df_edited = st.data_editor(df_cart, width="stretch", num_rows="dynamic", key="editor_data",
                    column_config={"SOLICITADA": st.column_config.NumberColumn("Solicitada", width="small"),
                                   "SURTIDO": st.column_config.NumberColumn("Surtido", width="small"),
                                   "O.C.": st.column_config.TextColumn("O.C.", width="small")})
                
                if not df_edited.equals(df_cart): st.session_state.carrito = df_edited.to_dict('records')
                
                # Callback para guardar pedido completo
                def finalizar_pedido_cb():
                    if st.session_state.cliente_box:
                        cod_cli, nom_cli = st.session_state.cliente_box.split(" - ", 1)
                        pedido_nuevo = {
                            "cli_cod": cod_cli,
                            "cli_nom": nom_cli,
                            "fecha": fecha_input,
                            "items": pd.DataFrame(st.session_state.carrito)
                        }
                        st.session_state.pedidos.append(pedido_nuevo)
                        st.session_state.carrito = []
                        st.session_state.cliente_box = None
                        st.session_state.search_faltantes_input = "" # Limpieza extra por si acaso
                    else:
                        st.error("Falta Cliente")

                st.button("💾 TERMINAR PEDIDO", type="primary", use_container_width=True, on_click=finalizar_pedido_cb)
            else:
                st.info("El carrito está vacío.")

    with tab2:
        st.metric("Pedidos Listos", len(st.session_state.pedidos))
        for i, p in enumerate(st.session_state.pedidos):
            with st.expander(f"{i+1}. {p['cli_nom']}"):
                st.dataframe(p['items'])
                if st.button("Borrar", key=f"del_{i}"):
                    st.session_state.pedidos.pop(i); st.rerun()
        
        if st.button("🚀 GENERAR EXCEL", disabled=(len(st.session_state.pedidos)==0)):
            try:
                wb = openpyxl.load_workbook(FILE_PLANTILLA)
                base = wb.active; base.title = "Base"
                conteo = {}
                
                for p in st.session_state.pedidos:
                    cod = p['cli_cod']
                    conteo[cod] = conteo.get(cod, 0) + 1
                    nom_hoja = cod if conteo[cod] == 1 else f"{cod}-{conteo[cod]}"
                    ws = wb.copy_worksheet(base); ws.title = nom_hoja
                    
                    ws['B2'] = "SUC. TIJ"; ws['B3'] = "LUIS FELIPE GARCÍA DOMÍNGUEZ"
                    ws['B4'] = p['cli_nom']; ws['D6'] = p['fecha'].strftime('%d/%m/%Y')
                    try: ws['B6'] = int(cod)
                    except: ws['B6'] = cod
                    
                    # Insertar Imagen (con manejo de errores si no existe)
                    try:
                        img = Image(FILE_IMAGEN)
                        img.width = 270; img.height = 80; img.anchor = 'D1'
                        ws.add_image(img)
                    except: pass
                    
                    datos = p['items'][['CODIGO', 'DESCRIPCION', 'SOLICITADA', 'SURTIDO', 'O.C.']].values.tolist()
                    for idx, row in enumerate(datos):
                        for c, val in enumerate(row): ws.cell(row=10+idx, column=c+1, value=val)
                
                del wb['Base']
                b = BytesIO(); wb.save(b); b.seek(0)
                st.download_button("⬇️ DESCARGAR", data=b, file_name="Faltantes.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            except Exception as e: st.error(f"Error: {e}")
