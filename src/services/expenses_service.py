import openpyxl
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO
from datetime import datetime
from src.config import FILE_PLANTILLA_COMPROBACION, FILE_PLANTILLA_SOLICITUD, FILE_IMAGEN, FILE_FIRMA

# Conversión aproximada: 1 cm ≈ 37.8 pixeles (96 DPI)
def cm_to_px(cm):
    return int(cm * 37.8)

def generar_excel_comprobacion(datos):
    """
    Genera el Excel de Comprobación de Viáticos.
    datos: Diccionario con llaves:
           - rango_fechas: 'DD/MM/AAAA - DD/MM/AAAA'
           - fecha_elaboracion: date
           - total_depositado: float
           - items: Lista de diccionarios {fecha, factura, concepto, importe, tasa, ish, observaciones}
    """
    wb = openpyxl.load_workbook(FILE_PLANTILLA_COMPROBACION)
    ws = wb.active

    # 1. IMAGEN (B1) - 1.4 cm
    try:
        img = XLImage(FILE_IMAGEN)
        # Altura 1.4cm, mantener aspecto
        # aspect ratio = w / h. New w = new h * ratio
        ratio = img.width / img.height
        h_px = cm_to_px(1.4)
        w_px = int(h_px * ratio)
        img.height = h_px
        img.width = w_px
        img.anchor = 'B1'
        ws.add_image(img)
    except Exception as e:
        print(f"Error imagen logo: {e}")

    # 2. HEADER
    ws['G5'] = datos['rango_fechas']
    
    # Formato de fecha personalizado: [nombre dia], [numero dia] de [mes] de [año]
    # Ejemplo: Sábado, 11 de Enero de 2026
    fecha_elab = datos['fecha_elaboracion']
    dias_semana = ["lunes", "martes", "miércoles", "jueves", "viernes", "sábado", "domingo"]
    meses_anio = ["enero", "febrero", "marzo", "abril", "mayo", "junio", "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"]
    
    dia_str = dias_semana[fecha_elab.weekday()]
    mes_str = meses_anio[fecha_elab.month - 1]
    fecha_fmt = f"{dia_str}, {fecha_elab.day} de {mes_str} de {fecha_elab.year}"
    
    cell_fecha = ws['C24']
    cell_fecha.value = fecha_fmt
    from openpyxl.styles import Font
    cell_fecha.font = Font(italic=True)
    
    ws['H25'] = datos['total_depositado']

    # 3. ITEMS (Filas 10 a 22)
    # B10: Fecha, C10: Factura, D10: Concepto, E10: Importe, F10: Tasa, G10: ISH, I10: Obs
    
    # Ordenar items por fecha (de menor a mayor)
    items_ordenados = sorted(datos['items'], key=lambda x: x['fecha'])
    
    fila_inicial = 10
    for i, item in enumerate(items_ordenados):
        if i >= 13: break # Solo hay 13 filas disponibles
        fila = fila_inicial + i
        
        ws[f'B{fila}'] = item['fecha'].strftime('%d/%m/%Y') if item.get('fecha') else ''
        ws[f'C{fila}'] = item.get('factura', '')
        
        concepto = item.get('concepto', '')
        if item.get('concepto_otro'): concepto = item.get('concepto_otro')
        ws[f'D{fila}'] = concepto
        
        ws[f'E{fila}'] = item.get('importe', 0)
        ws[f'F{fila}'] = item.get('tasa', 0)
        ws[f'G{fila}'] = item.get('ish', 0)
        
        obs = item.get('observaciones', '')
        if item.get('observaciones_otro'): obs = item.get('observaciones_otro')
        ws[f'I{fila}'] = obs

    b = BytesIO()
    wb.save(b)
    b.seek(0)
    return b

def generar_excel_solicitud(datos):
    """
    Genera el Excel de Solicitud de Viáticos.
    datos: Diccionario con llaves:
           - fecha_solicitud: date
           - inicio_periodo: date
           - fin_periodo: date
           - monto_solicitado: float
           - costo_estimado: float
           - presupuesto: {
               alimentos: {monto, dias},
               combustible: {monto, dias},
               hospedaje: {monto, dias},
               transporte: {monto, dias}
             }
    """
    wb = openpyxl.load_workbook(FILE_PLANTILLA_SOLICITUD)
    ws = wb.active

    # 1. LOGO (A1) - 1.35 cm
    try:
        img = XLImage(FILE_IMAGEN)
        ratio = img.width / img.height
        h_px = cm_to_px(1.35)
        w_px = int(h_px * ratio)
        img.height = h_px
        img.width = w_px
        img.anchor = 'A1'
        ws.add_image(img)
    except Exception as e:
        print(f"Error imagen logo: {e}")

    # 2. DATOS GENERALES
    ws['G4'] = datos['fecha_solicitud'].strftime('%d/%m/%Y')
    ws['C5'] = datos['inicio_periodo'].strftime('%d/%m/%Y')
    ws['C6'] = datos['fin_periodo'].strftime('%d/%m/%Y')
    ws['G5'] = datos['monto_solicitado']
    ws['C10'] = datos['costo_estimado']

    # 3. PRESUPUESTO
    p = datos['presupuesto']
    
    # Alimentos
    ws['D26'] = p['alimentos']['monto']
    ws['G26'] = p['alimentos']['dias']
    # Combustible
    ws['D28'] = p['combustible']['monto']
    ws['G28'] = p['combustible']['dias']
    # Hospedaje
    ws['D30'] = p['hospedaje']['monto']
    ws['G30'] = p['hospedaje']['dias']
    # Transporte
    ws['D32'] = p['transporte']['monto']
    ws['G32'] = p['transporte']['dias']

    # 4. FIRMA (E38) - 1.8 cm
    try:
        firma = XLImage(FILE_FIRMA)
        ratio_f = firma.width / firma.height
        h_f = cm_to_px(1.8)
        w_f = int(h_f * ratio_f)
        firma.height = h_f
        firma.width = w_f
        firma.anchor = 'E38'
        ws.add_image(firma)
    except Exception as e:
        print(f"Error imagen firma: {e}")

    b = BytesIO()
    wb.save(b)
    b.seek(0)
    return b


