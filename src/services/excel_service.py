import openpyxl
from io import BytesIO
from src.config import FILE_PLANTILLA

def generar_excel_faltantes(pedidos, fecha_elab, fecha_ini, fecha_fin):
    """
    Genera el archivo Excel a partir de la lista de pedidos usando el nuevo formato.
    Todos los faltantes se consolidan en una sola hoja.
    """
    wb = openpyxl.load_workbook(FILE_PLANTILLA)
    ws = wb.active
    ws.title = "Faltantes"
    
    # 1. ENCABEZADO - Celdas específicas
    # E7: Fecha de elaboración
    ws['E7'] = fecha_elab.strftime('%d/%m/%Y')
    
    # E8: Periodo (DD/MM/AAAA - DD/MM/AAAA)
    periodo_str = f"{fecha_ini.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"
    ws['E8'] = periodo_str
    
    # I7: Num Agente (Siempre "3354")
    ws['I7'] = "3354"
    
    # 2. CUERPO - A partir de la fila 11, columnas D a J
    # D: Sucursal ("TIJUANA")
    # E: Agente ("LUIS FELIPE GARCÍA DOMÍNGUEZ")
    # F: Código (Producto)
    # G: Descripción (Producto)
    # H: Cantidad Faltantes
    # I: Cliente (Código únicamente)
    # J: Orden de Compra ("NA")
    
    row_idx = 11
    
    # Consolidar todos los items de todos los pedidos
    for p in pedidos:
        cli_cod = p['cli_cod']
        df_items = p['items']
        
        for _, item in df_items.iterrows():
            ws.cell(row=row_idx, column=4, value="TIJUANA")                     # D
            ws.cell(row=row_idx, column=5, value="LUIS FELIPE GARCÍA DOMÍNGUEZ") # E
            ws.cell(row=row_idx, column=6, value=item['CODIGO'])               # F
            ws.cell(row=row_idx, column=7, value=item['DESCRIPCION'])           # G
            ws.cell(row=row_idx, column=8, value=item['SOLICITADA'])           # H
            ws.cell(row=row_idx, column=9, value=cli_cod)                       # I
            ws.cell(row=row_idx, column=10, value="NA")                         # J
            row_idx += 1
    
    # El logo ya no se agrega por petición del usuario
    
    b = BytesIO()
    wb.save(b)
    b.seek(0)
    return b
